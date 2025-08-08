import pandas as pd
from datetime import datetime
import os
import sys
import openpyxl

# Recebe o CNPJ e Ano_edit como argumentos
if len(sys.argv) < 2:
    print("Uso: python calculo_final.py <CNPJ> <Ano_edit>")
    sys.exit(1)
cnpj_input = sys.argv[1]
ano_edit_input = sys.argv[2]

#cnpj_input = "13.338.734/0001-27" 
#ano_edit_input = 2024

wb = openpyxl.load_workbook(r"U:\RBE\_PÚBLICO (Apenas trocas de arquivos)\Rating - Shared\pesos_financeiro.xlsm", data_only=True)
ws = wb['Planilha1']


db_sistema = pd.read_csv(r'Rating\db.csv')
db_financeiro = pd.read_csv(r'Financeiro/notas_financeiro.csv')
db_qualitativo = pd.read_csv(r'Qualitativo/notas_qualitativo.csv')
db_quantitativo = pd.read_csv(r'Quantitativo/notas_quantitativo.csv')

# Considere apenas linhas com status == "ativa"
if 'status' in db_financeiro.columns:
    db_financeiro = db_financeiro[db_financeiro['status'] == 'ativa']
if 'status' in db_qualitativo.columns:
    db_qualitativo = db_qualitativo[db_qualitativo['status'] == 'ativa']
if 'status' in db_quantitativo.columns:
    db_quantitativo = db_quantitativo[db_quantitativo['status'] == 'ativa']

resultados = []

# Filtra apenas o CNPJ e Ano_edit recebidos
df_cnpj = db_sistema[
    (db_sistema['CNPJ'].astype(str).str.strip() == str(cnpj_input).strip()) &
    (db_sistema['Ano_edit'].astype(str).str.strip() == str(ano_edit_input).strip())
]
for _, row in df_cnpj.iterrows():
    cnpj = row['CNPJ']
    ano_edit = row['Ano_edit']
    tipo_perfil = row.get('Tipo (Perfil Empresa)', '').strip().lower()

    # Busca as notas para o CNPJ e Ano_edit nos respectivos arquivos
    mask_fin = (db_financeiro['CNPJ'] == cnpj)
    if 'Ano_edit' in db_financeiro.columns:
        mask_fin = mask_fin & (db_financeiro['Ano_edit'] == ano_edit)
    nota_fin = db_financeiro.loc[mask_fin, 'Nota_Geral_Financeiro']

    mask_qual = (db_qualitativo['CNPJ'] == cnpj)
    if 'Ano_edit' in db_qualitativo.columns:
        mask_qual = mask_qual & (db_qualitativo['Ano_edit'] == ano_edit)
    nota_qual = db_qualitativo.loc[mask_qual, 'nota_final']

    mask_quant = (db_quantitativo['CNPJ'] == cnpj)
    if 'Ano_edit' in db_quantitativo.columns:
        mask_quant = mask_quant & (db_quantitativo['Ano_edit'] == ano_edit)
    nota_quant = db_quantitativo.loc[mask_quant, 'nota_final']

    nota_financeiro = nota_fin.values[0] if not nota_fin.empty else None
    nota_qualitativo = nota_qual.values[0] if not nota_qual.empty else None
    nota_quantitativo = nota_quant.values[0] if not nota_quant.empty else None

    # Definir pesos conforme o tipo de perfil
    if tipo_perfil == 'trading':
        peso_fin = float(ws['B19'].value)
        peso_qual = float(ws['B20'].value)
        peso_quant = float(ws['B21'].value)
    elif tipo_perfil == 'gerador':
        peso_fin = float(ws['C19'].value)
        peso_qual = float(ws['C20'].value)
        peso_quant = float(ws['C21'].value)
    elif tipo_perfil == 'consumidor':
        peso_fin = float(ws['D19'].value)
        peso_qual = float(ws['D20'].value)
        peso_quant = float(ws['D21'].value)
    elif tipo_perfil == 'fundo de investimento':
        peso_fin = float(ws['E19'].value)
        peso_qual = float(ws['E20'].value)
        peso_quant = float(ws['E21'].value)
    else:
        # padrão caso não reconheça o tipo
        peso_fin = 0.5
        peso_qual = 0.25
        peso_quant = 0.25

    # Calcular nota final conforme os pesos
    if None not in (nota_financeiro, nota_qualitativo, nota_quantitativo):
        nota_final = (
            float(nota_financeiro) * peso_fin +
            float(nota_qualitativo) * peso_qual +
            float(nota_quantitativo) * peso_quant
        )
    else:
        nota_final = None

    def calc_nota_final(valor):
        if valor is None:
            return None
        if valor < 1:
            return "D"
        elif valor < 2:
            return "CCC"
        elif valor < 2.5:
            return "B"
        elif valor < 3:
            return "BB"
        elif valor < 3.5:
            return "BBB"
        elif valor < 4:
            return "A"
        elif valor < 4.3:
            return "AA"
        else:
            return "AAA" 

    rating = calc_nota_final(nota_final)
    
    resultados.append({
        'CNPJ': cnpj,
        'Ano_edit': ano_edit,
        'nota_financeiro': nota_financeiro,
        'nota_qualitativo': nota_qualitativo,
        'nota_quantitativo': nota_quantitativo,
        'nota_final': nota_final,
        'rating': rating,
        'DataHora': datetime.now().strftime('%Y-%m-%d %H:%M'),
        'status': 'ativa'
    })

df_resultados = pd.DataFrame(resultados)
csv_path = 'Rating/notas_finais_gerais.csv'
if os.path.exists(csv_path):
    df_existente = pd.read_csv(csv_path)
    for _, row in df_resultados.iterrows():
        cnpj_norm = str(row['CNPJ']).replace('.', '').replace('/', '').replace('-', '').strip()
        ano_edit_norm = str(row['Ano_edit']).strip()
        mask = (
            df_existente['CNPJ'].astype(str).str.replace('.', '').str.replace('/', '').str.replace('-', '').str.strip() == cnpj_norm
        )
        if 'Ano_edit' in df_existente.columns:
            mask = mask & (df_existente['Ano_edit'].astype(str).str.strip() == ano_edit_norm)
        if mask.any():
            df_existente.loc[mask, 'status'] = 'editada'
    df_existente.to_csv(csv_path, index=False)
    df_resultados['status'] = 'ativa'
    df_resultados.to_csv(csv_path, mode='a', header=False, index=False)
else:
    df_resultados.to_csv(csv_path, index=False)