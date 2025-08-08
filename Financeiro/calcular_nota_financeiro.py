import pandas as pd
from datetime import datetime
import os
import sys
import openpyxl



# Recebe o CNPJ como argumento
if len(sys.argv) <= 2:
    print("Erro: forneça um CNPJ como argumento.")
    sys.exit(1)
cnpj_arg = sys.argv[1].strip()
ano_arg = sys.argv[2].strip()
db = pd.read_csv(r'Rating\db.csv')


wb = openpyxl.load_workbook(r"U:\RBE\_PÚBLICO (Apenas trocas de arquivos)\Rating - Shared\pesos_financeiro.xlsm", data_only=True)
ws = wb['Planilha1']

# Filtra apenas o CNPJ e Ano_edit fornecidos
db['CNPJ'] = db['CNPJ'].astype(str).str.strip()
if 'Ano_edit' in db.columns:
    db['Ano_edit'] = db['Ano_edit'].astype(str).str.strip()
    db = db[(db['CNPJ'] == cnpj_arg) & (db['Ano_edit'] == ano_arg)]
else:
    db = db[db['CNPJ'] == cnpj_arg]
    db['Ano_edit'] = ano_arg  # adiciona coluna se não existir

# Considere apenas linhas com status == "ativa"
if 'status' in db.columns:
    db = db[db['status'] == 'ativa']

# Função para converter valores para float, tratando vírgulas e valores ausentes
def to_float(x):
    try:
        if pd.isna(x):
            return 0.0
        return float(str(x).replace(',', '.'))
    except:
        return 0.0

# region Ranges para notas

def nota_liquidez_corrente_trading(valor):
    # Exemplo: ajuste os intervalos conforme a tabela da imagem para 'trading'
    if valor < 0.5:
        return -2
    elif valor < 0.65:
        return -1
    elif valor < 0.8:
        return 0
    elif valor < 0.9:
        return 1
    elif valor < 1.0:
        return 2
    elif valor < 1.25:
        return 3
    elif valor < 1.5:
        return 4
    else:
        return 5

def nota_liquidez_corrente_gerador(valor):
    # Exemplo: ajuste os intervalos conforme a tabela da imagem para 'gerador'
    if valor < 0.6:
        return -2
    elif valor < 0.7:
        return -1
    elif valor < 0.8:
        return 0
    elif valor < 0.9:
        return 1
    elif valor < 1:
        return 2
    elif valor < 1.10:
        return 3
    elif valor < 1.15:
        return 4
    else:
        return 5

def nota_liquidez_corrente_consumidor(valor):
    # Exemplo: ajuste os intervalos conforme a tabela da imagem para 'consumidor'
    if valor < 0.60:
        return -2
    elif valor < 0.7:
        return -1
    elif valor < 0.80:
        return 0
    elif valor < 1.0:
        return 1
    elif valor < 1.10:
        return 2
    elif valor < 1.30:
        return 3
    elif valor < 1.5:
        return 4
    else:
        return 5

def nota_liquidez_corrente_fundo_investimento(valor):
    # Exemplo: ajuste os intervalos conforme a tabela da imagem para 'fundo de investimento'
    if valor < 0.6:
        return -2
    elif valor < 0.7:
        return -1
    elif valor < 0.8:
        return 0
    elif valor < 1:
        return 1
    elif valor < 1.10:
        return 2
    elif valor < 1.3:
        return 3
    elif valor < 1.5:
        return 4
    else:
        return 5

def nota_liquidez_corrente_padrao(valor):
    # Caso não seja nenhum dos tipos conhecidos
    if valor < 0.5:
        return -2
    elif valor < 0.65:
        return -1
    elif valor < 0.8:
        return 0
    elif valor < 0.9:
        return 1
    elif valor < 1.0:
        return 2
    elif valor < 1.25:
        return 3
    elif valor < 1.5:
        return 4
    else:
        return 5

def nota_liquidez_corrente_por_tipo(valor, tipo):
    if isinstance(tipo, str):
        tipo = tipo.strip().lower()
    else:
        tipo = ""
    if tipo == 'trading':
        return nota_liquidez_corrente_trading(valor)
    elif tipo == 'gerador':
        return nota_liquidez_corrente_gerador(valor)
    elif tipo == 'consumidor':
        return nota_liquidez_corrente_consumidor(valor)
    elif tipo == 'fundo de investimento':
        return nota_liquidez_corrente_fundo_investimento(valor)
    # Adicione outros tipos conforme necessário
    else:
        return nota_liquidez_corrente_padrao(valor)

def nota_liquidez_geral_trading(valor):
    if valor < 0.15:
        return -2
    elif valor < 0.3:
        return -1
    elif valor < 0.6:
        return 0
    elif valor < 0.9:
        return 1
    elif valor < 1.05:
        return 2
    elif valor < 1.15:
        return 3
    elif valor < 1.25:
        return 4
    else:
        return 5

def nota_liquidez_geral_gerador(valor):
    if valor < 0.15:
        return -2
    elif valor < 0.3:
        return -1
    elif valor < 0.6:
        return 0
    elif valor < 0.9:
        return 1
    elif valor < 1.05:
        return 2
    elif valor < 1.15:
        return 3
    elif valor < 1.25:
        return 4
    else:
        return 5

def nota_liquidez_geral_consumidor(valor):
    if valor < 0.15:
        return -2
    elif valor < 0.3:
        return -1
    elif valor < 0.6:
        return 0
    elif valor < 0.9:
        return 1
    elif valor < 1.05:
        return 2
    elif valor < 1.15:
        return 3
    elif valor < 1.25:
        return 4
    else:
        return 5

def nota_liquidez_geral_fundo_investimento(valor):
    if valor < 0.15:
        return -2
    elif valor < 0.3:
        return -1
    elif valor < 0.6:
        return 0
    elif valor < 0.9:
        return 1
    elif valor < 1.05:
        return 2
    elif valor < 1.15:
        return 3
    elif valor < 1.25:
        return 4
    else:
        return 5

def nota_liquidez_geral_padrao(valor):
    if valor < 0.15:
        return -2
    elif valor < 0.3:
        return -1
    elif valor < 0.6:
        return 0
    elif valor < 0.9:
        return 1
    elif valor < 1.05:
        return 2
    elif valor < 1.15:
        return 3
    elif valor < 1.25:
        return 4
    else:
        return 5

def nota_liquidez_geral_por_tipo(valor, tipo):
    if isinstance(tipo, str):
        tipo = tipo.strip().lower()
    else:
        tipo = ""
    if tipo == 'trading':
        return nota_liquidez_geral_trading(valor)
    elif tipo == 'gerador':
        return nota_liquidez_geral_gerador(valor)
    elif tipo == 'consumidor':
        return nota_liquidez_geral_consumidor(valor)
    elif tipo == 'fundo de investimento':
        return nota_liquidez_geral_fundo_investimento(valor)
    else:
        return nota_liquidez_geral_padrao(valor)

def nota_patrimonio_liquido_trading(valor):
    if valor < -5000000:
        return -2
    elif valor < 0:
        return -1
    elif valor < 5000000:
        return 0
    elif valor < 10000000:
        return 1
    elif valor < 20000000:
        return 2
    elif valor < 40000000:
        return 3
    elif valor < 80000000:
        return 4
    else:
        return 5

def nota_patrimonio_liquido_gerador(valor):
    if valor < -5000000:
        return -2
    elif valor < 0:
        return -1
    elif valor < 5000000:
        return 0
    elif valor < 10000000:
        return 1
    elif valor < 20000000:
        return 2
    elif valor < 40000000:
        return 3
    elif valor < 80000000:
        return 4
    else:
        return 5

def nota_patrimonio_liquido_consumidor(valor):
    if valor < -5000000:
        return -2
    elif valor < 0:
        return -1
    elif valor < 1000000:
        return 0
    elif valor < 5000000:
        return 1
    elif valor < 1000000:
        return 2
    elif valor < 30000000:
        return 3
    elif valor < 50000000:
        return 4
    else:
        return 5

def nota_patrimonio_liquido_fundo_investimento(valor):
    if valor < -5000000:
        return -2
    elif valor < 0:
        return -1
    elif valor < 5000000:
        return 0
    elif valor < 10000000:
        return 1
    elif valor < 20000000:
        return 2
    elif valor < 40000000:
        return 3
    elif valor < 80000000:
        return 4
    else:
        return 5

def nota_patrimonio_liquido_padrao(valor):
    if valor < -5000000:
        return -2
    elif valor < 0:
        return -1
    elif valor < 5000000:
        return 0
    elif valor < 10000000:
        return 1
    elif valor < 20000000:
        return 2
    elif valor < 40000000:
        return 3
    elif valor < 80000000:
        return 4
    else:
        return 5

def nota_patrimonio_liquido_por_tipo(valor, tipo):
    if isinstance(tipo, str):
        tipo = tipo.strip().lower()
    else:
        tipo = ""
    if tipo == 'trading':
        return nota_patrimonio_liquido_trading(valor)
    elif tipo == 'gerador':
        return nota_patrimonio_liquido_gerador(valor)
    elif tipo == 'consumidor':
        return nota_patrimonio_liquido_consumidor(valor)
    elif tipo == 'fundo de investimento':
        return nota_patrimonio_liquido_fundo_investimento(valor)
    else:
        return nota_patrimonio_liquido_padrao(valor)

def nota_termometro_insolvencia_trading(valor):
    if valor < -5:
        return -2
    elif valor < -1:
        return -1
    elif valor < 0:
        return 0
    elif valor < 1:
        return 1
    elif valor < 2:
        return 2
    elif valor < 4:
        return 3
    elif valor < 7:
        return 4
    else:
        return 5

def nota_termometro_insolvencia_gerador(valor):
    if valor < -5:
        return -2
    elif valor < -1:
        return -1
    elif valor < 0:
        return 0
    elif valor < 1:
        return 1
    elif valor < 2:
        return 2
    elif valor < 4:
        return 3
    elif valor < 7:
        return 4
    else:
        return 5

def nota_termometro_insolvencia_consumidor(valor):
    if valor < -5:
        return -2
    elif valor < -1:
        return -1
    elif valor < 0:
        return 0
    elif valor < 1:
        return 1
    elif valor < 2:
        return 2
    elif valor < 4:
        return 3
    elif valor < 7:
        return 4
    else:
        return 5

def nota_termometro_insolvencia_fundo_investimento(valor):
    if valor < -5:
        return -2
    elif valor < -1:
        return -1
    elif valor < 0:
        return 0
    elif valor < 1:
        return 1
    elif valor < 2:
        return 2
    elif valor < 4:
        return 3
    elif valor < 7:
        return 4
    else:
        return 5

def nota_termometro_insolvencia_padrao(valor):
    if valor < -5:
        return -2
    elif valor < -1:
        return -1
    elif valor < 0:
        return 0
    elif valor < 1:
        return 1
    elif valor < 2:
        return 2
    elif valor < 4:
        return 3
    elif valor < 7:
        return 4
    else:
        return 5

def nota_termometro_insolvencia_por_tipo(valor, tipo):
    if isinstance(tipo, str):
        tipo = tipo.strip().lower()
    else:
        tipo = ""
    if tipo == 'trading':
        return nota_termometro_insolvencia_trading(valor)
    elif tipo == 'gerador':
        return nota_termometro_insolvencia_gerador(valor)
    elif tipo == 'consumidor':
        return nota_termometro_insolvencia_consumidor(valor)
    elif tipo == 'fundo de investimento':
        return nota_termometro_insolvencia_fundo_investimento(valor)
    else:
        return nota_termometro_insolvencia_padrao(valor)

def nota_endividamento_terceiros_trading(valor):
    if valor > 1:
        return -2
    elif valor > 0.9:
        return -1
    elif valor > 0.8:
        return 0
    elif valor > 0.6:
        return 1
    elif valor > 0.4:
        return 2
    elif valor > 0.25:
        return 3
    elif valor > 0.1:
        return 4
    else:
        return 5

def nota_endividamento_terceiros_gerador(valor):
    if valor > 3:
        return -2
    elif valor > 2:
        return -1
    elif valor > 1.75:
        return 0
    elif valor > 1.5:
        return 1
    elif valor > 1.25:
        return 2
    elif valor > 1:
        return 3
    elif valor > 0.75:
        return 4
    else:
        return 5

def nota_endividamento_terceiros_consumidor(valor):
    if valor > 1.5:
        return -2
    elif valor > 1.2:
        return -1
    elif valor > 1:
        return 0
    elif valor > 0.8:
        return 1
    elif valor > 0.6:
        return 2
    elif valor > 0.35:
        return 3
    elif valor > 0.25:
        return 4
    else:
        return 5

def nota_endividamento_terceiros_fundo_investimento(valor):
    if valor > 1.5:
        return -2
    elif valor > 1.25:
        return -1
    elif valor > 1:
        return 0
    elif valor > 0.75:
        return 1
    elif valor > 0.5:
        return 2
    elif valor > 0.3:
        return 3
    elif valor > 0.2:
        return 4
    else:
        return 5

def nota_endividamento_terceiros_padrao(valor):
    if valor > 1:
        return -2
    elif valor > 0.9:
        return -1
    elif valor > 0.8:
        return 0
    elif valor > 0.6:
        return 1
    elif valor > 0.4:
        return 2
    elif valor > 0.25:
        return 3
    elif valor > 0.1:
        return 4
    else:
        return 5

def nota_endividamento_terceiros_por_tipo(valor, tipo):
    if isinstance(tipo, str):
        tipo = tipo.strip().lower()
    else:
        tipo = ""
    if tipo == 'trading':
        return nota_endividamento_terceiros_trading(valor)
    elif tipo == 'gerador':
        return nota_endividamento_terceiros_gerador(valor)
    elif tipo == 'consumidor':
        return nota_endividamento_terceiros_consumidor(valor)
    elif tipo == 'fundo de investimento':
        return nota_endividamento_terceiros_fundo_investimento(valor)
    else:
        return nota_endividamento_terceiros_padrao(valor)

def nota_gro_trading(valor):
    if valor > 1:
        return -2
    elif valor > 0.9:
        return -1
    elif valor > 0.8:
        return 0
    elif valor > 0.6:
        return 1
    elif valor > 0.3:
        return 2
    elif valor > 0.15:
        return 3
    elif valor > 0.05:
        return 4
    else:
        return 5

def nota_gro_gerador(valor):
    if valor > 1:
        return -2
    elif valor > 0.9:
        return -1
    elif valor > 0.8:
        return 0
    elif valor > 0.4:
        return 1
    elif valor > 0.3:
        return 2
    elif valor > 0.2:
        return 3
    elif valor > 0.1:
        return 4
    else:
        return 5

def nota_gro_consumidor(valor):
    if valor > 1:
        return -2
    elif valor > 0.9:
        return -1
    elif valor > 0.8:
        return 0
    elif valor > 0.4:
        return 1
    elif valor > 0.3:
        return 2
    elif valor > 0.2:
        return 3
    elif valor > 0.1:
        return 4
    else:
        return 5

def nota_gro_fundo_investimento(valor):
    if valor > 1:
        return -2
    elif valor > 0.9:
        return -1
    elif valor > 0.8:
        return 0
    elif valor > 0.6:
        return 1
    elif valor > 0.3:
        return 2
    elif valor > 0.2:
        return 3
    elif valor > 0.1:
        return 4
    else:
        return 5

def nota_gro_padrao(valor):
    if valor > 1:
        return -2
    elif valor > 0.9:
        return -1
    elif valor > 0.8:
        return 0
    elif valor > 0.4:
        return 1
    elif valor > 0.3:
        return 2
    elif valor > 0.2:
        return 3
    elif valor > 0.1:
        return 4
    else:
        return 5

def nota_gro_por_tipo(valor, tipo):
    if isinstance(tipo, str):
        tipo = tipo.strip().lower()
    else:
        tipo = ""
    if tipo == 'trading':
        return nota_gro_trading(valor)
    elif tipo == 'gerador':
        return nota_gro_gerador(valor)
    elif tipo == 'consumidor':
        return nota_gro_consumidor(valor)
    elif tipo == 'fundo de investimento':
        return nota_gro_fundo_investimento(valor)
    else:
        return nota_gro_padrao(valor)

def nota_caixa_equivalentes_trading(valor):
    if valor < 0:
        return -2
    elif valor < 1000000:
        return -1
    elif valor < 3000000:
        return 0
    elif valor < 5000000:
        return 1
    elif valor < 20000000:
        return 2
    elif valor < 40000000:
        return 3
    elif valor < 70000000:
        return 4
    else:
        return 5

def nota_caixa_equivalentes_gerador(valor):
    if valor < 0:
        return -2
    elif valor < 1000000:
        return -1
    elif valor < 3000000:
        return 0
    elif valor < 5000000:
        return 1
    elif valor < 15000000:
        return 2
    elif valor < 30000000:
        return 3
    elif valor < 60000000:
        return 4
    else:
        return 5

def nota_caixa_equivalentes_consumidor(valor):
    if valor < 0:
        return -2
    elif valor < 1000000:
        return -1
    elif valor < 5000000:
        return 0
    elif valor < 10000000:
        return 1
    elif valor < 25000000:
        return 2
    elif valor < 50000000:
        return 3
    elif valor < 100000000:
        return 4
    else:
        return 5

def nota_caixa_equivalentes_fundo_investimento(valor):
    if valor < 0:
        return -2
    elif valor < 1000000:
        return -1
    elif valor < 5000000:
        return 0
    elif valor < 10000000:
        return 1
    elif valor < 25000000:
        return 2
    elif valor < 50000000:
        return 3
    elif valor < 100000000:
        return 4
    else:
        return 5

def nota_caixa_equivalentes_padrao(valor):
    if valor < 0:
        return -2
    elif valor < 1000000:
        return -1
    elif valor < 5000000:
        return 0
    elif valor < 10000000:
        return 1
    elif valor < 25000000:
        return 2
    elif valor < 50000000:
        return 3
    elif valor < 100000000:
        return 4
    else:
        return 5

def nota_caixa_equivalentes_por_tipo(valor, tipo):
    if isinstance(tipo, str):
        tipo = tipo.strip().lower()
    else:
        tipo = ""
    if tipo == 'trading':
        return nota_caixa_equivalentes_trading(valor)
    elif tipo == 'gerador':
        return nota_caixa_equivalentes_gerador(valor)
    elif tipo == 'consumidor':
        return nota_caixa_equivalentes_consumidor(valor)
    elif tipo == 'fundo de investimento':
        return nota_caixa_equivalentes_fundo_investimento(valor)
    else:
        return nota_caixa_equivalentes_padrao(valor)

def nota_auditoria_func(valor):
    if valor == "":
        return 0
    else:
        return 5
    
def nota_pmp_func(valor):
    if valor > 100:
        return -2
    elif valor > 90:
        return -1
    elif valor > 60:
        return 0
    elif valor > 45:
        return 1
    elif valor > 30:
        return 2
    elif valor > 15:
        return 3
    elif valor > 5:
        return 4
    else:
        return 5
    
def nota_pmr_func(valor):
    if valor > 100:
        return -2
    elif valor > 90:
        return -1
    elif valor > 60:
        return 0
    elif valor > 45:
        return 1
    elif valor > 30:
        return 2
    elif valor > 15:
        return 3
    elif valor > 5:
        return 4
    else:
        return 5


def nota_alavancagem_ebitda_trading(valor):
    if valor > 5:
        return -2
    elif valor > 4.5:
        return -1
    elif valor > 4:
        return 0
    elif valor > 3.5:
        return 1
    elif valor > 3:
        return 2
    elif valor > 2:
        return 3
    elif valor > 1:
        return 4
    else:
        return 5


def nota_alavancagem_ebitda_gerador(valor):
    if valor > 5:
        return -2
    elif valor > 4.5:
        return -1
    elif valor > 4:
        return 0
    elif valor > 3.5:
        return 1
    elif valor > 3:
        return 2
    elif valor > 2:
        return 3
    elif valor > 1:
        return 4
    else:
        return 5


def nota_alavancagem_ebitda_consumidor(valor):
    if valor > 5:
        return -2
    elif valor > 4.5:
        return -1
    elif valor > 4:
        return 0
    elif valor > 3.5:
        return 1
    elif valor > 3:
        return 2
    elif valor > 2:
        return 3
    elif valor > 1:
        return 4
    else:
        return 5


def nota_alavancagem_ebitda_fundo_investimento(valor):
    if valor > 5:
        return -2
    elif valor > 4.5:
        return -1
    elif valor > 4:
        return 0
    elif valor > 3.5:
        return 1
    elif valor > 3:
        return 2
    elif valor > 2:
        return 3
    elif valor > 1:
        return 4
    else:
        return 5


def nota_alavancagem_ebitda_padrao(valor):
    if valor > 5:
        return -2
    elif valor > 4.5:
        return -1
    elif valor > 4:
        return 0
    elif valor > 3.5:
        return 1
    elif valor > 3:
        return 2
    elif valor > 2:
        return 3
    elif valor > 1:
        return 4
    else:
        return 5

def nota_alavancagem_ebitda_por_tipo(valor, tipo):
    if isinstance(tipo, str):
        tipo = tipo.strip().lower()
    else:
        tipo = ""
    if tipo == 'trading':
        return nota_alavancagem_ebitda_trading(valor)
    elif tipo == 'gerador':
        return nota_alavancagem_ebitda_gerador(valor)
    elif tipo == 'consumidor':
        return nota_alavancagem_ebitda_consumidor(valor)
    elif tipo == 'fundo de investimento':
        return nota_alavancagem_ebitda_fundo_investimento(valor)
    else:
        return nota_alavancagem_ebitda_padrao(valor)




resultados = []

# endregion  

for idx, row in db.iterrows():

    #---------------------------- -----------------------------------
    #                         BUSCAR POR COLUNAS NO CSV 
    #---------------------------- -----------------------------------
    cnpj = row.get('CNPJ', '')
    ano = row.get('Ano_edit', '')
    tipo_empresa = row.get('Tipo (Perfil Empresa)', '')
    ativo_circ = to_float(row.get('Ativo Circulante', 0))
    passivo_circ = to_float(row.get('Passivo Circulante', 0))
    outras_receber_lp = to_float(row.get('Outras Contas Receber LP', 0))
    passivo_nc = to_float(row.get('Passivo Não Circulante', 0))
    capital_social = to_float(row.get('Capital Social', 0))
    reserva_lucros = to_float(row.get('Reserva de Lucros', 0))
    outros_pl = to_float(row.get('Outros Itens do PL', 0))
    lucro_liquido = to_float(row.get('Lucro Líquido', 0))
    estoques = to_float(row.get('Estoques', 0))
    cpv = to_float(row.get('CPV', 0))
    despesas_operacionais = to_float(row.get('Despesas Operacionais', 0))
    clientes = to_float(row.get('Clientes', 0))
    fornecedores = to_float(row.get('Fornecedores', 0))
    emprestimos_cp = to_float(row.get('Empréstimos CP', 0))
    emprestimos_lp = to_float(row.get('Empréstimos LP', 0))
    #
    #---------ROL,CAIXA_EQ e AUDITORIA - se encontram já na aba calculos
    #
    #---------------------------- -----------------------------------
    #                               CALCULOS
    #---------------------------- -----------------------------------
    #1
    liquidez_corrente = ativo_circ / passivo_circ if passivo_circ != 0 else 0
    #2
    liquidez_geral = (ativo_circ + outras_receber_lp) / (passivo_circ + passivo_nc) if (passivo_circ + passivo_nc) != 0 else 0
    #3
    patrimonio_liquido = capital_social + reserva_lucros + outros_pl
    #4
    liquidez_seca = (ativo_circ - estoques) / passivo_circ if passivo_circ != 0 else 0
    endividamento_terceiros = (passivo_circ + passivo_nc) / abs(patrimonio_liquido) if patrimonio_liquido != 0 else 0
    termometro_insolvencia = (
        (lucro_liquido / patrimonio_liquido * 0.05 if patrimonio_liquido != 0 else 0)
        + (liquidez_geral * 1.65)
        + (liquidez_seca * 3.55)
        - (liquidez_corrente * 1.06)
        - (endividamento_terceiros * 0.33)
    )
    #5
    endividamento_terceiros = endividamento_terceiros
    #6
    rol = to_float(row.get('Receita Operacional Líquida(ROL)',0))
    gro = rol / patrimonio_liquido if patrimonio_liquido != 0 else 0
    #7
    caixa_eq = to_float(row.get('Caixa e Equivalente de caixa', 0))
    #8
    auditoria = to_float(row.get('Auditoria', 0))
    #9
    pmp = clientes / rol * 365 if rol != 0 else 101
    #10
    pmr = fornecedores / abs(cpv) * 365 if cpv != 0 else 101
    #11
    alavancagem_ebidta = (passivo_circ + passivo_nc - caixa_eq) / (rol + cpv + despesas_operacionais) if (rol + cpv + despesas_operacionais) != 0 else 101
    #nc na nota
    alavancagem_cp = emprestimos_cp / (emprestimos_cp + emprestimos_lp) if (emprestimos_cp + emprestimos_lp) != 0 else 0
    #---------------------------- -----------------------------------
    #                            NOTAS 
    #---------------------------- -----------------------------------
    nota_liq_corrente = nota_liquidez_corrente_por_tipo(liquidez_corrente, tipo_empresa)
    nota_liq_geral = nota_liquidez_geral_por_tipo(liquidez_geral, tipo_empresa)
    nota_patrimonio_liquido = nota_patrimonio_liquido_por_tipo(patrimonio_liquido, tipo_empresa)  
    nota_endividamento_terceiros = nota_endividamento_terceiros_por_tipo(endividamento_terceiros, tipo_empresa)
    nota_gro = nota_gro_por_tipo(gro, tipo_empresa)
    nota_termometro_insolvencia = nota_termometro_insolvencia_por_tipo(termometro_insolvencia, tipo_empresa)
    nota_caixa_eq = nota_caixa_equivalentes_por_tipo(caixa_eq, tipo_empresa)
    nota_auditoria = nota_auditoria_func(auditoria)
    nota_pmp = nota_pmp_func(pmp)
    nota_pmr = nota_pmr_func(pmr)
    nota_alavancagem_ebitda = nota_alavancagem_ebitda_por_tipo(alavancagem_ebidta, tipo_empresa)
    #---------------------------- -----------------------------------
    #                     PONDERAÇÃO POR CONTRAPARTE 
    #---------------------------- -----------------------------------
    tipo = tipo_empresa.strip().lower() if isinstance(tipo_empresa, str) else ""

    if tipo == 'trading':
        pesos = {
            'liq_corrente': float(ws['B2'].value),
            'liq_geral': float(ws['B3'].value),
            'patrimonio_liquido': float(ws['B4'].value),
            'termometro_insolvencia': float(ws['B5'].value),
            'endividamento_terceiros': float(ws['B6'].value),
            'gro': float(ws['B7'].value),
            'caixa_eq': float(ws['B8'].value),
            'auditoria' : float(ws['B9'].value),
            'pmp' : float(ws['B10'].value),
            'pmr' : float(ws['B11'].value),
            'alavancagem_ebitda' : float(ws['B12'].value)
        }
    elif tipo == 'gerador':
        pesos = {
            'liq_corrente': float(ws['C2'].value),
            'liq_geral': float(ws['C3'].value),
            'patrimonio_liquido': float(ws['C4'].value),
            'termometro_insolvencia': float(ws['C5'].value),
            'endividamento_terceiros': float(ws['C6'].value),
            'gro': float(ws['C7'].value),
            'caixa_eq': float(ws['C8'].value),
            'auditoria' : float(ws['C9'].value),
            'pmp' : float(ws['C10'].value),
            'pmr' : float(ws['C11'].value),
            'alavancagem_ebitda' : float(ws['C12'].value)
        }
    elif tipo == 'consumidor':
        pesos = {
            'liq_corrente': float(ws['D2'].value),
            'liq_geral': float(ws['D3'].value),
            'patrimonio_liquido': float(ws['D4'].value),
            'termometro_insolvencia': float(ws['D5'].value),
            'endividamento_terceiros': float(ws['D6'].value),
            'gro': float(ws['D7'].value),
            'caixa_eq': float(ws['D8'].value),
            'auditoria' : float(ws['D9'].value),
            'pmp' : float(ws['D10'].value),
            'pmr' : float(ws['D11'].value),
            'alavancagem_ebitda' : float(ws['D12'].value)       
        }
    elif tipo == 'fundo de investimento':
        pesos = {
            'liq_corrente': float(ws['E2'].value),
            'liq_geral': float(ws['E3'].value),
            'patrimonio_liquido': float(ws['E4'].value),
            'termometro_insolvencia': float(ws['E5'].value),
            'endividamento_terceiros': float(ws['E6'].value),
            'gro': float(ws['E7'].value),
            'caixa_eq': float(ws['E8'].value),
            'auditoria' : float(ws['E9'].value),
            'pmp' : float(ws['E10'].value),
            'pmr' : float(ws['E11'].value),
            'alavancagem_ebitda' : float(ws['E12'].value)      
        }
    else:
        pesos = {
            'liq_corrente': float(ws['B2'].value),
            'liq_geral': float(ws['B3'].value),
            'patrimonio_liquido': float(ws['B4'].value),
            'termometro_insolvencia': float(ws['B5'].value),
            'endividamento_terceiros': float(ws['B6'].value),
            'gro': float(ws['B7'].value),
            'caixa_eq': float(ws['B8'].value),
            'auditoria' : float(ws['B9'].value),
            'pmp' : float(ws['B10'].value),
            'pmr' : float(ws['B11'].value),
            'alavancagem_ebitda' : float(ws['B12'].value)
        }

    #---------------------------------------------------------------
    #                   SOMA DAS NOTAS PARA NOTA FINAL  
    #---------------------------------------------------------------
    nota_geral_financeiro = (
        nota_liq_corrente * pesos['liq_corrente'] +
        nota_liq_geral * pesos['liq_geral'] +
        nota_patrimonio_liquido * pesos['patrimonio_liquido'] +
        nota_termometro_insolvencia * pesos['termometro_insolvencia'] +
        nota_endividamento_terceiros * pesos['endividamento_terceiros'] +
        nota_gro * pesos['gro'] +
        nota_caixa_eq * pesos['caixa_eq'] +
        nota_auditoria * pesos['auditoria'] +
        nota_pmp * pesos['pmp'] + 
        nota_pmr * pesos['pmr'] +
        nota_alavancagem_ebitda * pesos['alavancagem_ebitda']  
    )


    resultados.append({
        'CNPJ': cnpj,
        'Ano_edit': ano_arg,
        'Nota_Liquidez_Corrente': nota_liq_corrente,
        'Nota_Liquidez_Geral': nota_liq_geral,
        'Nota_Patrimonio_Liquido': nota_patrimonio_liquido,
        'Nota_Termometro_de_Insolvencia': nota_termometro_insolvencia,
        'Nota_Estrutura_de_Capital': nota_endividamento_terceiros,
        'Nota_Gro': nota_gro,
        'Nota_Caixa_Equivalentes_de_Caixa': nota_caixa_eq,
        'Nota_Auditoria' : nota_auditoria,
        'Nota_Pmp' : nota_pmp,
        'Nota_Pmr' : nota_pmr,
        'Nota_Alavancagem_Ebitda' : nota_alavancagem_ebitda,
        'Alavancagem_CP': alavancagem_cp,
        'Nota_Geral_Financeiro': nota_geral_financeiro,
        'DataHora': datetime.now().strftime('%Y-%m-%d %H:%M'),
        'status': 'ativa'
    })

df_resultados = pd.DataFrame(resultados)
csv_path = r'Financeiro\notas_financeiro.csv'
if os.path.exists(csv_path):
    df_existente = pd.read_csv(csv_path)
    for idx, row in df_resultados.iterrows():
        mask = (df_existente['CNPJ'].astype(str).str.strip() == str(row['CNPJ']).strip()) & \
               (df_existente['Ano_edit'].astype(str).str.strip() == str(row['Ano_edit']).strip())
        if mask.any():
            df_existente.loc[mask, 'status'] = 'editada'
    df_existente.to_csv(csv_path, index=False)
    df_resultados.to_csv(csv_path, mode='a', header=False, index=False)
else:
    df_resultados.to_csv(csv_path, index=False)